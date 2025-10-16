# -*- coding: utf-8 -*-
# vaa_korea_mapped_report.py
import os
from datetime import datetime
import pandas as pd
import yfinance as yf


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

# =========================================================
# 설정
# =========================================================
CONVERT_TO_KRW = True   # 미국 ETF 가격을 원화로 환산해 표기 (USDKRW=X 월말환율)
MIN_MONTHS = 13         # 12개월 비교(현재 포함)에 필요한 최소 월 스냅샷 수

OUT_DIR = "vaa_out"

# =========================================================
# 의사결정(미국 ETF, 백데이터) 구성
# =========================================================
DECISION_ASSETS = [
    # 분류, 의사결정기준(라벨), 미국ETF 티커
    ["공격자산", "미국 주식SPY", "SPY"],
    ["공격자산", "선진국 주식EFA", "EFA"],
    ["공격자산", "개발도상국 주식EEM", "EEM"],
    ["공격자산", "미국 혼합채권AGG", "AGG"],
    ["안전자산", "미국 회사채LQD", "LQD"],
    ["안전자산", "미국 중기국채IEF", "IEF"],
    ["안전자산", "미국 단기국채SHY", "SHY"],
]
DECISION_DF = pd.DataFrame(DECISION_ASSETS, columns=["분류", "의사결정기준", "US_Ticker"])

# =========================================================
# 실제 투자(국내 ETF, ISA) 매핑: 미국ETF → 국내ETF
# =========================================================
US_TO_KR_MAP = {
    "SPY":  {"종목명": "KODEX 미국S&P500",              "Code": "379800", "환율": "환노출"},
    "EFA":  {"종목명": "KODEX MSCI선진국",              "Code": "251350", "환율": "환노출"},
    "EEM":  {"종목명": "KODEX MSCI EM선물(H)",          "Code": "291890", "환율": "환해지"},
    "AGG":  {"종목명": "KODEX iShares미국투자등급회사채 엑티브", "Code": "468630", "환율": "환노출"},
    "LQD":  {"종목명": "KODEX 미국종합채권ESG엑티브(H)", "Code": "437080", "환율": "환해지"},
    "IEF":  {"종목명": "ACE 미국10년국채엑티브",         "Code": "0085P0", "환율": "환노출"},
    "SHY":  {"종목명": "ACE 미국달러단기채권엑티브",     "Code": "440650", "환율": "환노출"},
}

# (옵션) 프록시 – 특정 미국 ETF가 데이터 부족/이상일 때 대체
PROXY_MAP = {
    "SPY": ["IVV", "VOO"],
    "EFA": ["IEFA"],
    "EEM": ["VWO"],
    "AGG": ["BND"],
    "LQD": ["VCIT"],
    "IEF": ["GOVT"],
    "SHY": ["BIL", "SHV"],
}

# =========================================================
# 환율 (KRW per USD) 월말 시리즈
# =========================================================
def get_usdkrw_series(start="2010-01-01") -> pd.Series:
    fx = yf.download("USDKRW=X", start=start, progress=False)
    if fx.empty or "Close" not in fx:
        raise RuntimeError("환율(USDKRW=X) 데이터를 불러오지 못했습니다.")
    return fx["Close"].resample("M").last().dropna()

USDKRW_MONTHLY = get_usdkrw_series()

# =========================================================
# 데이터 핸들링
# =========================================================
def load_daily(ticker: str, start="2010-01-01") -> pd.DataFrame:
    return yf.download(ticker, start=start, progress=False)

def monthly_with_current(df: pd.DataFrame) -> pd.Series:
    """
    일봉 → 월말 종가 + 현재(오늘) 종가 보강.
    CONVERT_TO_KRW=True 이면 월별 환율(USDKRW=X) 곱해 KRW Series로 변환.
    항상 float Series 반환.
    """
    if df is None or df.empty or "Close" not in df:
        return pd.Series(dtype="float64")

    close = df["Close"]
    if isinstance(close, pd.DataFrame):
        close = close.iloc[:, 0]

    monthly = close.resample("M").last().dropna()

    # 현재 종가 보강
    last_date = close.index[-1]
    last_close = close.iloc[-1]
    if len(monthly) == 0 or monthly.index[-1].month != last_date.month or monthly.index[-1].year != last_date.year:
        monthly = pd.concat([monthly, pd.Series([last_close], index=[last_date])])

    # KRW 환산
    if CONVERT_TO_KRW:
        fx = USDKRW_MONTHLY.reindex(monthly.index, method="ffill")
        if isinstance(fx, pd.DataFrame):
            fx = fx.iloc[:, 0]
        monthly = monthly.astype("float64") * fx.astype("float64")

    if isinstance(monthly, pd.DataFrame):
        monthly = monthly.squeeze("columns")
    return monthly.astype("float64")

def snapshot_momentum(monthly: pd.Series):
    """
    스냅샷 모멘텀: r_k = P0/Pk - 1
    score_raw = 12*r1 + 4*r3 + 2*r6 + 1*r12   (비율)
    score_pct = score_raw * 100                (퍼센트)
    반환: r1,r3,r6,r12(비율), score_pct(퍼센트), P0,P1,P3,P6,P12(가격)
    """
    if monthly is None or len(monthly) < MIN_MONTHS:
        raise ValueError("월말 시리즈가 부족합니다.")

    def _scalar(x):
        if hasattr(x, "item"):
            try: return float(x.item())
            except Exception: pass
        try: return float(x)
        except Exception:
            if isinstance(x, pd.Series):
                return float(x.iloc[0])
            raise

    P0  = _scalar(monthly.iloc[-1])   # 현재
    P1  = _scalar(monthly.iloc[-2])   # 1개월 전
    P3  = _scalar(monthly.iloc[-4])   # 3개월 전
    P6  = _scalar(monthly.iloc[-7])   # 6개월 전
    P12 = _scalar(monthly.iloc[-13])  # 12개월 전

    r1  = P0 / P1  - 1.0
    r3  = P0 / P3  - 1.0
    r6  = P0 / P6  - 1.0
    r12 = P0 / P12 - 1.0

    score_raw = 12*r1 + 4*r3 + 2*r6 + 1*r12
    score_pct = score_raw * 100.0

    return r1, r3, r6, r12, score_pct, P0, P1, P3, P6, P12

def resolve_with_proxy(us_ticker: str):
    """
    미국 ETF tiicker로 월 스냅샷 생성.
    필요시 PROXY_MAP을 이용해 대체.
    """
    df = load_daily(us_ticker)
    monthly = monthly_with_current(df)
    if len(monthly) >= MIN_MONTHS:
        return us_ticker, "원본", *snapshot_momentum(monthly)

    for p in PROXY_MAP.get(us_ticker, []):
        dfp = load_daily(p)
        monthly_p = monthly_with_current(dfp)
        if len(monthly_p) >= MIN_MONTHS:
            return p, f"대체[{p}]", *snapshot_momentum(monthly_p)

    return us_ticker, "데이터없음", None, None, None, None, None, None, None, None, None, None

# =========================================================
# 계산 & 결정
# =========================================================
def build_summary_df():
    """
    Summary DataFrame:
    [분류, 의사결정기준, US_Ticker, 사용티커, 데이터출처,
     실제투자_종목명, 실제투자_Code, 실제투자_환율,
     1/3/6/12개월(%), 모멘텀점수(가중합,%),
     현재/1/3/6/12개월 가격(KRW)]
    """
    rows = []
    for _, r in DECISION_DF.iterrows():
        group, label, us_ticker = r["분류"], r["의사결정기준"], r["US_Ticker"]

        used_ticker, src, r1, r3, r6, r12, score_pct, P0, P1, P3, P6, P12 = resolve_with_proxy(us_ticker)

        # 국내 투자 종목 매핑
        kr_map = US_TO_KR_MAP.get(us_ticker, {"종목명": None, "Code": None, "환율": None})

        rows.append([
            group, label, us_ticker, used_ticker, src,
            kr_map["종목명"], kr_map["Code"], kr_map["환율"],
            None if r1  is None else round(r1*100,  2),
            None if r3  is None else round(r3*100,  2),
            None if r6  is None else round(r6*100,  2),
            None if r12 is None else round(r12*100, 2),
            None if score_pct is None else round(score_pct, 2),
            P0, P1, P3, P6, P12
        ])

    return pd.DataFrame(rows, columns=[
        "분류","의사결정기준","US_Ticker","사용티커","데이터출처",
        "실제투자_종목명","실제투자_Code","실제투자_환율",
        "1개월(%)","3개월(%)","6개월(%)","12개월(%)","모멘텀점수(가중합,%)",
        "현재가격(KRW)","1개월전(KRW)","3개월전(KRW)","6개월전(KRW)","12개월전(KRW)"
    ])

def decision_banner(summary_df: pd.DataFrame) -> str:
    """
    규칙:
    - 공격 4개(SPY,EFA,EEM,AGG)의 모멘텀점수 모두 > 0 → 그중 최고점의 국내 ETF에 투자
    - 아니면 안전(LQD, IEF, SHY)에서 최고점의 국내 ETF에 투자
    """
    aggr = summary_df[summary_df["분류"]=="공격자산"].copy()
    safe = summary_df[summary_df["분류"]=="안전자산"].copy()

    if all((aggr["모멘텀점수(가중합,%)"] > 0).fillna(False)):
        tgt = aggr.loc[aggr["모멘텀점수(가중합,%)"].idxmax()]
    else:
        tgt = safe.loc[safe["모멘텀점수(가중합,%)"].idxmax()]

    # 배너: 국내 실제 투자 종목과 코드 표시 + (의사결정기준)
    return f"이번달 투자 대상: {tgt['실제투자_종목명']} ({tgt['실제투자_Code']})  —  기준: {tgt['의사결정기준']} / {tgt['US_Ticker']}"

# =========================================================
# 엑셀 생성
# =========================================================
def autosize_columns(ws, max_width=48):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            v = "" if v is None else str(v)
            widths[i] = max(widths.get(i, 0), len(v))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), max_width)

def write_summary_sheet(wb: Workbook, df: pd.DataFrame, month_str: str):
    ws = wb.create_sheet("Summary")
    title_fill = PatternFill("solid", fgColor="E6F0FF")
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="D9D9D9")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Styles
    if "percent_style" not in wb.named_styles:
        st = NamedStyle(name="percent_style"); st.number_format = "0.00%"; wb.add_named_style(st)
    if "won_style" not in wb.named_styles:
        st = NamedStyle(name="won_style"); st.number_format = '#,##0"원"'; wb.add_named_style(st)

    # Banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    c = ws.cell(row=1, column=1, value=f"VAA Summary — {month_str} (가격단위: {'KRW' if CONVERT_TO_KRW else 'USD'})")
    c.font = Font(size=14, bold=True); c.fill = title_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Header
    for col, h in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True); cell.fill = header_fill
        cell.border = border_all; cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data
    start_row = 3
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_all
            header = df.columns[c_idx-1]
            if header.endswith("(%)") and isinstance(val, (int, float)):
                cell.value = val / 100.0
                cell.style = "percent_style"
            if header.endswith("(KRW)") and isinstance(val, (int, float)):
                cell.style = "won_style"

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(df.columns))}{ws.max_row}"
    autosize_columns(ws, max_width=44)

def write_detail_sheet(wb: Workbook, df: pd.DataFrame, banner: str, month_str: str):
    ws = wb.create_sheet("Detail")
    title_fill = PatternFill("solid", fgColor="E6F0FF")
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="D9D9D9")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    if "percent_style" not in wb.named_styles:
        st = NamedStyle(name="percent_style"); st.number_format = "0.00%"; wb.add_named_style(st)
    if "won_style" not in wb.named_styles:
        st = NamedStyle(name="won_style"); st.number_format = '#,##0"원"'; wb.add_named_style(st)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    c = ws.cell(row=1, column=1, value=f"VAA Detail — {month_str} (가격단위: {'KRW' if CONVERT_TO_KRW else 'USD'})")
    c.font = Font(size=14, bold=True); c.fill = title_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.cell(row=3, column=1, value=banner).font = Font(bold=True)
    row_cursor = 5

    def add_block(ar):
        nonlocal row_cursor
        # 상단 정보
        ws.cell(row=row_cursor, column=1, value="의사결정기준"); ws.cell(row=row_cursor, column=2, value=f"{ar['의사결정기준']} / {ar['US_Ticker']}"); row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="실제 투자");   ws.cell(row=row_cursor, column=2, value=f"{ar['실제투자_종목명']} ({ar['실제투자_Code']})"); row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="환율표기");   ws.cell(row=row_cursor, column=2, value=ar["실제투자_환율"]); row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="모멘텀 스코어")
        sc = ws.cell(row=row_cursor, column=2, value=(ar["모멘텀점수(가중합,%)"]/100.0 if pd.notna(ar["모멘텀점수(가중합,%)"]) else None))
        sc.number_format = "0.00%"; row_cursor += 1
        row_cursor += 1

        # 표 헤더
        headers = ["구간","현재","1개월 전","3개월 전","6개월 전","12개월 전"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_cursor, column=col, value=h)
            cell.font = Font(bold=True); cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center"); cell.border = border_all
        row_cursor += 1

        # 가격
        price_row = ["가격", ar["현재가격(KRW)"], ar["1개월전(KRW)"], ar["3개월전(KRW)"], ar["6개월전(KRW)"], ar["12개월전(KRW)"]]
        for col, v in enumerate(price_row, start=1):
            cell = ws.cell(row=row_cursor, column=col, value=v)
            cell.border = border_all
            if col > 1 and isinstance(v, (int, float)): cell.style = "won_style"
        row_cursor += 1

        # 수익률
        r_row = ["각 구간 수익률", None, ar["1개월(%)"], ar["3개월(%)"], ar["6개월(%)"], ar["12개월(%)"]]
        for col, v in enumerate(r_row, start=1):
            if col <= 2:
                cell = ws.cell(row=row_cursor, column=col, value=v)
            else:
                cell = ws.cell(row=row_cursor, column=col, value=None if pd.isna(v) else v/100.0)
                if not pd.isna(v): cell.number_format = "0.00%"
            cell.border = border_all
        row_cursor += 1

        # 배수
        mult_row = ["각 구간 배수", None, 12, 4, 2, 1]
        for col, v in enumerate(mult_row, start=1):
            cell = ws.cell(row=row_cursor, column=col, value=v); cell.border = border_all
        row_cursor += 1

        # 각 스코어(비율→% 표시)
        s1  = None if pd.isna(ar["1개월(%)"]) else ar["1개월(%)"] * 12 / 100.0
        s3  = None if pd.isna(ar["3개월(%)"]) else ar["3개월(%)"] * 4  / 100.0
        s6  = None if pd.isna(ar["6개월(%)"]) else ar["6개월(%)"] * 2  / 100.0
        s12 = None if pd.isna(ar["12개월(%)"]) else ar["12개월(%)"] * 1  / 100.0
        s_row = ["각 스코어", None, s1, s3, s6, s12]
        for col, v in enumerate(s_row, start=1):
            cell = ws.cell(row=row_cursor, column=col, value=v)
            cell.border = border_all
            if col > 2 and v is not None: cell.number_format = "0.00%"
        row_cursor += 1

        # 합계
        for col in range(1, 6):
            cell = ws.cell(row=row_cursor, column=col, value=""); cell.border = border_all
        tot = ws.cell(row=row_cursor, column=6, value=(ar["모멘텀점수(가중합,%)"]/100.0 if pd.notna(ar["모멘텀점수(가중합,%)"]) else None))
        tot.number_format = "0.00%"; tot.border = border_all
        row_cursor += 2

    # 공격 → 안전 순서대로 출력
    for grp in ["공격자산","안전자산"]:
        sub = df[df["분류"]==grp]
        if sub.empty: continue
        ws.cell(row=row_cursor, column=1, value=grp).font = Font(size=12, bold=True)
        row_cursor += 1
        for _, ar in sub.iterrows():
            add_block(ar)

    ws.freeze_panes = "A5"
    autosize_columns(ws, max_width=60)

# =========================================================
# 메인
# =========================================================
def main():
    month_str = datetime.now().strftime("%Y-%m")
    os.makedirs(OUT_DIR, exist_ok=True)
    xlsx_path = os.path.join(OUT_DIR, f"vaa_report_{month_str}.xlsx")

    summary = build_summary_df()
    banner  = decision_banner(summary)

    wb = Workbook(); wb.remove(wb.active)
    write_summary_sheet(wb, summary, month_str)
    write_detail_sheet(wb, summary, banner, month_str)
    wb.save(xlsx_path)

    print(f"✅ 엑셀 저장 완료: {xlsx_path}")
    print(f"📌 {banner}")

if __name__ == "__main__":
    main()
