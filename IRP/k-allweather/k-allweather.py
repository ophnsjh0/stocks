# -*- coding: utf-8 -*-
import math
from datetime import datetime
import pandas as pd
import FinanceDataReader as fdr

# ---- 포트폴리오 정의 (성장형) ----
# 비율은 0~1 사이로 표기
ASSETS = [
    {"종목명": "KODEX 미국 S&P500TR",          "종목코드": "379800", "비율": 0.24},
    {"종목명": "KOSEF 200TR",                   "종목코드": "294400", "비율": 0.08},
    {"종목명": "KODEX 차이나CSI300",            "종목코드": "283580", "비율": 0.08},
    {"종목명": "KODEX 인도 Nifty50",            "종목코드": "453810", "비율": 0.08},
    {"종목명": "ACE KRX금현물",                 "종목코드": "411060", "비율": 0.19},
    {"종목명": "KODEX 미국10년국채선물",          "종목코드": "308620", "비율": 0.07},
    {"종목명": "ACE 미국30년국채액티브(H)",     "종목코드": "453850", "비율": 0.07},
    {"종목명": "KBSTAR KIS 국고채 30년 Enhanced","종목코드": "385560", "비율": 0.14},
    {"종목명": "TIGER KOFR금리액티브(합성)",     "종목코드": "449170", "비율": 0.05},
]

def get_last_price(krx_code: str) -> float:
    """
    FinanceDataReader에서 KRX 종목코드의 최근 종가를 반환.
    장중에는 당일 데이터가 갱신되지 않았을 수 있음(이 경우 전일 종가 사용).
    """
    df = fdr.DataReader(krx_code)  # KRX는 숫자코드 문자열 그대로 사용
    if df is None or df.empty:
        raise RuntimeError(f"가격 조회 실패: {krx_code}")
    return float(df["Close"].iloc[-1])

def build_allocation(total_krw: int) -> pd.DataFrame:
    rows = []
    for a in ASSETS:
        price = get_last_price(a["종목코드"])
        target_amt = total_krw * a["비율"]
        qty = math.floor(target_amt / price)  # 정수 주 구매
        buy_amt = qty * price
        rows.append({
            "종목명": a["종목명"],
            "종목코드": a["종목코드"],
            "%비율": a["비율"],
            "현재가": price,
            "목표금액": target_amt,
            "보유수량": qty,
            "실제매수금액": buy_amt,
            "잔여(목표-실제)": target_amt - buy_amt,
        })
    df = pd.DataFrame(rows)
    df.loc["합계", ["%비율","현재가"]] = [df["%비율"].sum(), None]
    df.loc["합계","목표금액"] = df["목표금액"].sum()
    df.loc["합계","실제매수금액"] = df["실제매수금액"].sum()
    df.loc["합계","잔여(목표-실제)"] = df["잔여(목표-실제)"].sum()
    df["투자금액"] = df["목표금액"]  # 요청 컬럼명에 맞춰 복제
    return df

def save_to_excel(df: pd.DataFrame, total_krw: int, path: str):
    # 보기 좋은 형식으로 저장
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment, Font, NamedStyle

    wb = Workbook()
    ws = wb.active
    ws.title = "K-올웨더(성장형)"

    # 제목
    title = f"K-올웨더 (성장형) 배분표 - 투자금액: {total_krw:,} KRW - 생성일 {datetime.now():%Y-%m-%d %H:%M}"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # 데이터
    out_cols = ["종목명","종목코드","%비율","투자금액","현재가","보유수량","실제매수금액","잔여(목표-실제)"]
    ws.append(out_cols)
    for r in dataframe_to_rows(df[out_cols], index=False, header=False):
        ws.append(r)

    # 서식
    pct = NamedStyle(name="pct"); pct.number_format = "0.00%"
    krw = NamedStyle(name="krw"); krw.number_format = '#,##0'
    if "pct" not in wb.named_styles: wb.add_named_style(pct)
    if "krw" not in wb.named_styles: wb.add_named_style(krw)

    for row in ws.iter_rows(min_row=3, min_col=3, max_col=3, max_row=ws.max_row-1):
        for cell in row: cell.style = "pct"
    for col in [4,5,7,8]:
        for cell in ws.iter_rows(min_row=3, min_col=col, max_col=col, max_row=ws.max_row):
            for c in cell: c.style = "krw"

    # 열 너비
    widths = [34,12,8,16,12,10,16,16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    # 총 투자 대비 미집행 현금(잔여 합계)
    leftover = float(df.loc["합계","잔여(목표-실제)"])
    ws.append([])
    ws.append(["미집행 현금(잔여 합계)", leftover])
    ws["B{}".format(ws.max_row)].style = "krw"

    wb.save(path)

if __name__ == "__main__":
    # ▶ 투자금액 입력
    total = int(input("총 투자금액(KRW)을 입력하세요 (예: 10000000): ").strip())
    df = build_allocation(total)
    out_path = f"result/K_AllWeather_{datetime.now():%Y%m%d_%H%M}.xlsx"
    save_to_excel(df, total, out_path)
    print(f"✅ 엑셀 저장 완료: {out_path}")
