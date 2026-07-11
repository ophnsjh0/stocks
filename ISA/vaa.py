# -*- coding: utf-8 -*-
# vaa_korea_mapped_composite_report.py
# VAA 전략: EFA 대신 [유로스탁스50 + 일본니케이225] 합성 지수 사용 버전

import os
from datetime import datetime
import pandas as pd
import numpy as np
import FinanceDataReader as fdr

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

# =========================================================
# 설정
# =========================================================
CONVERT_TO_KRW = True   # 미국 ETF 가격을 원화로 환산해 표기 (단, 국내 ETF 합성은 제외)
MIN_MONTHS = 13         # 12개월 비교(현재 포함)에 필요한 최소 월 스냅샷 수

OUT_DIR = "vaa_out"

# =========================================================
# 의사결정 자산 구성 (EFA -> COMPOSITE_EU_JP 변경)
# =========================================================
DECISION_ASSETS = [
    # 분류, 의사결정기준(라벨), 티커(Key)
    ["공격자산", "미국 주식SPY", "SPY"],
    ["공격자산", "선진국(유로+일본)", "COMPOSITE_EU_JP"], # ✅ EFA 대체
    ["공격자산", "개발도상국 주식EEM", "EEM"],
    ["공격자산", "미국 혼합채권AGG", "AGG"],
    ["안전자산", "미국 회사채LQD", "LQD"],
    ["안전자산", "미국 중기국채IEF", "IEF"],
    ["안전자산", "미국 단기국채SHY", "SHY"],
]
DECISION_DF = pd.DataFrame(DECISION_ASSETS, columns=["분류", "의사결정기준", "US_Ticker"])

# =========================================================
# 실제 투자(국내 ETF) 매핑
# =========================================================
US_TO_KR_MAP = {
    "SPY":  {"종목명": "KODEX 미국S&P500",              "Code": "379800", "환율": "환노출"},
    # ✅ 합성 지수 매핑 정보 (엑셀 표기용)
    "COMPOSITE_EU_JP": {"종목명": "TIGER 유로50 / 일본225 (각 50%)", "Code": "195930 / 241180", "환율": "국내상장(KRW)"},
    "EEM":  {"종목명": "PLUS 신흥국MSCI(합성 H)",          "Code": "195980", "환율": "환해지"},
    "AGG":  {"종목명": "KODEX 미국종합채권ESG엑티브(H)", "Code": "437080", "환율": "환해지"},
    "LQD":  {"종목명": "KODEX iShares미국투자등급회사채 엑티브", "Code": "468630", "환율": "환노출"},
    "IEF":  {"종목명": "ACE 미국10년국채엑티브",         "Code": "0085P0", "환율": "환노출"},
    "SHY":  {"종목명": "ACE 미국달러단기채권엑티브",     "Code": "440650", "환율": "환노출"},
}

# (옵션) 프록시 – 특정 미국 ETF가 데이터 부족/이상일 때 대체
PROXY_MAP = {
    "SPY": ["IVV", "VOO"],
    # EFA 프록시는 제거 (합성 로직 사용)
    "EEM": ["VWO"],
    "AGG": ["BND"],
    "LQD": ["VCIT"],
    "IEF": ["GOVT"],
    "SHY": ["BIL", "SHV"],
}

# =========================================================
# 환율 (KRW per USD) 월말 시리즈
# =========================================================
def _read_fx_usdkrw(start="2010-01-01") -> pd.Series:
    candidates = ["USD/KRW", "USDKRW", "USD-KRW"]
    last_err = None
    for sym in candidates:
        try:
            fx = fdr.DataReader(sym, start)
            if not fx.empty and "Close" in fx:
                s = fx["Close"].dropna()
                return s.resample("M").last().dropna()
        except Exception as e:
            last_err = e
            continue
    print(f"⚠️ 환율 데이터 로드 실패 (마지막 오류: {last_err}) - 환산 기능을 끕니다.")
    global CONVERT_TO_KRW
    CONVERT_TO_KRW = False
    return pd.Series(dtype=float)

USDKRW_MONTHLY = _read_fx_usdkrw()

# =========================================================
# 데이터 핸들링 유틸리티
# =========================================================
def load_daily(ticker: str, start="2010-01-01") -> pd.DataFrame:
    try:
        df = fdr.DataReader(ticker, start)
        if df is None or df.empty or "Close" not in df.columns:
            return pd.DataFrame()
        if not isinstance(df.index, pd.DatetimeIndex):
            df.index = pd.to_datetime(df.index)
        df = df.sort_index()
        return df
    except Exception:
        return pd.DataFrame()

# ✅ [신규 함수] 유로+일본 합성 지수 생성기
def load_composite_daily(start="2010-01-01") -> pd.DataFrame:
    print(">> 합성 지수(유로+일본) 데이터 생성 중...")
    try:
        # TIGER 유로스탁스50(195930), TIGER 일본니케이225(241180)
        df_eu = fdr.DataReader("195930", start)
        df_jp = fdr.DataReader("241180", start)

        if df_eu.empty or df_jp.empty:
            return pd.DataFrame()

        # 종가 시리즈 추출
        s_eu = df_eu['Close']
        s_jp = df_jp['Close']

        # 날짜 교집합 (데이터 정렬)
        idx = s_eu.index.intersection(s_jp.index)
        s_eu = s_eu.loc[idx]
        s_jp = s_jp.loc[idx]

        # 일간 변동률 계산
        ret_eu = s_eu.pct_change().fillna(0)
        ret_jp = s_jp.pct_change().fillna(0)

        # 50:50 합성 수익률
        ret_composite = (ret_eu * 0.5) + (ret_jp * 0.5)

        # 지수화 (기준일 1,000pt 시작 가정)
        composite_idx = (1 + ret_composite).cumprod() * 1000.0
        
        # DataFrame 형태로 반환 (fdr 포맷과 맞춤)
        return pd.DataFrame({"Close": composite_idx}, index=idx)

    except Exception as e:
        print(f"❌ 합성 지수 생성 실패: {e}")
        return pd.DataFrame()

def monthly_with_current(df: pd.DataFrame, is_krw_asset: bool = False) -> pd.Series:
    """
    일봉 → 월말 종가 + 현재(오늘) 종가 보강.
    is_krw_asset=True이면 환율 곱셈을 건너뜀 (이미 KRW).
    """
    if df is None or df.empty or "Close" not in df:
        return pd.Series(dtype="float64")

    close = df["Close"]
    if isinstance(close, pd.DataFrame):
        close = close.iloc[:, 0]

    monthly = close.resample("M").last().dropna()

    # 현재 종가 보강
    last_date = close.index[-1]
    last_close = float(close.iloc[-1])
    if len(monthly) == 0 or (monthly.index[-1].month != last_date.month or monthly.index[-1].year != last_date.year):
        monthly = pd.concat([monthly, pd.Series([last_close], index=[last_date])])

    # KRW 환산 (이미 원화자산이 아니고, 변환 옵션이 켜져있을 때만)
    if CONVERT_TO_KRW and not is_krw_asset:
        fx = USDKRW_MONTHLY.reindex(monthly.index, method="ffill")
        if isinstance(fx, pd.DataFrame):
            fx = fx.iloc[:, 0]
        monthly = monthly.astype("float64") * fx.astype("float64")

    if isinstance(monthly, pd.DataFrame):
        monthly = monthly.squeeze("columns")
    return monthly.astype("float64")

def snapshot_momentum(monthly: pd.Series):
    """
    스냅샷 모멘텀 계산
    """
    if monthly is None or len(monthly) < MIN_MONTHS:
        raise ValueError("월말 시리즈가 부족합니다.")

    def _scalar(x):
        try: return float(x.item())
        except: return float(x)

    P0  = _scalar(monthly.iloc[-1])   # 현재
    P1  = _scalar(monthly.iloc[-2])   # 1개월 전
    P3  = _scalar(monthly.iloc[-4])   # 3개월 전
    P6  = _scalar(monthly.iloc[-7])   # 6개월 전
    P12 = _scalar(monthly.iloc[-13])  # 12개월 전

    r1  = P0 / P1  - 1.0
    r3  = P0 / P3  - 1.0
    r6  = P0 / P6  - 1.0
    r12 = P0 / P12 - 1.0

    score_raw = 12*r1 + 4*r3 + 2*r6 + 1*r12
    score_pct = score_raw # (기존 코드 로직 유지)

    return r1, r3, r6, r12, score_pct, P0, P1, P3, P6, P12

def resolve_with_proxy(ticker_key: str):
    """
    티커 키에 따라 데이터를 로드.
    - COMPOSITE_EU_JP: 합성 데이터 로드 (환율 적용 X)
    - 그 외: 미국 ETF 로드 (환율 적용 O)
    """
    # 1. 특수 케이스: 유로+일본 합성
    if ticker_key == "COMPOSITE_EU_JP":
        df = load_composite_daily()
        # 합성 지수는 이미 KRW 기반이므로 is_krw_asset=True
        monthly = monthly_with_current(df, is_krw_asset=True)
        if len(monthly) >= MIN_MONTHS:
            return ticker_key, "합성(KRW)", *snapshot_momentum(monthly)
        else:
            return ticker_key, "데이터부족", None, None, None, None, None, None, None, None, None, None

    # 2. 일반 케이스 (미국 ETF)
    # 원본 티커 시도
    df = load_daily(ticker_key)
    monthly = monthly_with_current(df, is_krw_asset=False)
    if len(monthly) >= MIN_MONTHS:
        return ticker_key, "원본", *snapshot_momentum(monthly)

    # 프록시 시도
    for p in PROXY_MAP.get(ticker_key, []):
        dfp = load_daily(p)
        monthly_p = monthly_with_current(dfp, is_krw_asset=False)
        if len(monthly_p) >= MIN_MONTHS:
            return p, f"대체[{p}]", *snapshot_momentum(monthly_p)

    return ticker_key, "데이터없음", None, None, None, None, None, None, None, None, None, None

# =========================================================
# 계산 & 결정
# =========================================================
def build_summary_df():
    rows = []
    for _, r in DECISION_DF.iterrows():
        group, label, us_ticker = r["분류"], r["의사결정기준"], r["US_Ticker"]

        used_ticker, src, r1, r3, r6, r12, score_pct, P0, P1, P3, P6, P12 = resolve_with_proxy(us_ticker)

        # 국내 투자 종목 매핑
        kr_map = US_TO_KR_MAP.get(us_ticker, {"종목명": None, "Code": None, "환율": None})

        rows.append([
            group, label, us_ticker, used_ticker, src,
            kr_map["종목명"], kr_map["Code"], kr_map["환율"],
            None if r1  is None else round(r1*100,  2),
            None if r3  is None else round(r3*100,  2),
            None if r6  is None else round(r6*100,  2),
            None if r12 is None else round(r12*100, 2),
            None if score_pct is None else round(score_pct, 2),
            P0, P1, P3, P6, P12
        ])

    return pd.DataFrame(rows, columns=[
        "분류","의사결정기준","US_Ticker","사용티커","데이터출처",
        "실제투자_종목명","실제투자_Code","실제투자_환율",
        "1개월(%)","3개월(%)","6개월(%)","12개월(%)","모멘텀점수(가중합,%)",
        "현재가격(KRW)","1개월전(KRW)","3개월전(KRW)","6개월전(KRW)","12개월전(KRW)"
    ])

def decision_banner(summary_df: pd.DataFrame) -> str:
    aggr = summary_df[summary_df["분류"]=="공격자산"].copy()
    safe = summary_df[summary_df["분류"]=="안전자산"].copy()

    # 공격자산 4개가 모두 모멘텀 > 0 인지 체크
    if all((aggr["모멘텀점수(가중합,%)"] > 0).fillna(False)):
        tgt = aggr.loc[aggr["모멘텀점수(가중합,%)"].idxmax()]
    else:
        tgt = safe.loc[safe["모멘텀점수(가중합,%)"].idxmax()]

    return f"이번달 투자 대상: {tgt['실제투자_종목명']} ({tgt['실제투자_Code']})  —  기준: {tgt['의사결정기준']} / {tgt['US_Ticker']}"

# =========================================================
# 엑셀 생성
# =========================================================
def autosize_columns(ws, max_width=48):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            v = "" if v is None else str(v)
            widths[i] = max(widths.get(i, 0), len(v))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), max_width)

def write_summary_sheet(wb: Workbook, df: pd.DataFrame, month_str: str):
    ws = wb.create_sheet("Summary")
    title_fill = PatternFill("solid", fgColor="E6F0FF")
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="D9D9D9")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    if "percent_style" not in wb.named_styles:
        st = NamedStyle(name="percent_style"); st.number_format = "0.00%"; wb.add_named_style(st)
    if "won_style" not in wb.named_styles:
        st = NamedStyle(name="won_style"); st.number_format = '#,##0"원"'; wb.add_named_style(st)

    # Banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    c = ws.cell(row=1, column=1, value=f"VAA Summary (EFA대체: 유로/일본 합성) — {month_str}")
    c.font = Font(size=14, bold=True); c.fill = title_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Header
    for col, h in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True); cell.fill = header_fill
        cell.border = border_all; cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data
    start_row = 3
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_all
            header = df.columns[c_idx-1]
            if header.endswith("(%)") and isinstance(val, (int, float)):
                cell.value = val / 100.0
                cell.style = "percent_style"
            if header.endswith("(KRW)") and isinstance(val, (int, float)):
                cell.style = "won_style"

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(df.columns))}{ws.max_row}"
    autosize_columns(ws, max_width=44)

def write_detail_sheet(wb: Workbook, df: pd.DataFrame, banner: str, month_str: str):
    ws = wb.create_sheet("Detail")
    title_fill = PatternFill("solid", fgColor="E6F0FF")
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="D9D9D9")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    if "percent_style" not in wb.named_styles:
        st = NamedStyle(name="percent_style"); st.number_format = "0.00%"; wb.add_named_style(st)
    if "won_style" not in wb.named_styles:
        st = NamedStyle(name="won_style"); st.number_format = '#,##0"원"'; wb.add_named_style(st)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    c = ws.cell(row=1, column=1, value=f"VAA Detail (가격단위: KRW) — {month_str}")
    c.font = Font(size=14, bold=True); c.fill = title_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.cell(row=3, column=1, value=banner).font = Font(bold=True)
    row_cursor = 5

    def add_block(ar):
        nonlocal row_cursor
        ws.cell(row=row_cursor, column=1, value="의사결정기준"); ws.cell(row=row_cursor, column=2, value=f"{ar['의사결정기준']} / {ar['US_Ticker']}"); row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="실제 투자");   ws.cell(row=row_cursor, column=2, value=f"{ar['실제투자_종목명']} ({ar['실제투자_Code']})"); row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="환율표기");   ws.cell(row=row_cursor, column=2, value=ar["실제투자_환율"]); row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="모멘텀 스코어")
        sc = ws.cell(row=row_cursor, column=2, value=(ar["모멘텀점수(가중합,%)"] if pd.notna(ar["모멘텀점수(가중합,%)"]) else None))
        sc.number_format = "0.00"; row_cursor += 1
        row_cursor += 1

        headers = ["구간","현재","1개월 전","3개월 전","6개월 전","12개월 전"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_cursor, column=col, value=h)
            cell.font = Font(bold=True); cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center"); cell.border = border_all
        row_cursor += 1

        price_row = ["가격", ar["현재가격(KRW)"], ar["1개월전(KRW)"], ar["3개월전(KRW)"], ar["6개월전(KRW)"], ar["12개월전(KRW)"]]
        for col, v in enumerate(price_row, start=1):
            cell = ws.cell(row=row_cursor, column=col, value=v)
            cell.border = border_all
            if col > 1 and isinstance(v, (int, float)): cell.style = "won_style"
        row_cursor += 1

        r_row = ["각 구간 수익률", None, ar["1개월(%)"], ar["3개월(%)"], ar["6개월(%)"], ar["12개월(%)"]]
        for col, v in enumerate(r_row, start=1):
            if col <= 2:
                cell = ws.cell(row=row_cursor, column=col, value=v)
            else:
                cell = ws.cell(row=row_cursor, column=col, value=None if pd.isna(v) else v/100.0)
                if not pd.isna(v): cell.number_format = "0.00%"
            cell.border = border_all
        row_cursor += 1

        mult_row = ["각 구간 배수", None, 12, 4, 2, 1]
        for col, v in enumerate(mult_row, start=1):
            cell = ws.cell(row=row_cursor, column=col, value=v); cell.border = border_all
        row_cursor += 1

        s1  = None if pd.isna(ar["1개월(%)"]) else ar["1개월(%)"] * 12 / 100.0
        s3  = None if pd.isna(ar["3개월(%)"]) else ar["3개월(%)"] * 4  / 100.0
        s6  = None if pd.isna(ar["6개월(%)"]) else ar["6개월(%)"] * 2  / 100.0
        s12 = None if pd.isna(ar["12개월(%)"]) else ar["12개월(%)"] * 1  / 100.0
        s_row = ["각 스코어", None, s1, s3, s6, s12]
        for col, v in enumerate(s_row, start=1):
            cell = ws.cell(row=row_cursor, column=col, value=v)
            cell.border = border_all
            if col > 2 and v is not None: cell.number_format = "0.00"
        row_cursor += 1

        for col in range(1, 6):
            cell = ws.cell(row=row_cursor, column=col, value=""); cell.border = border_all
        tot = ws.cell(row=row_cursor, column=6, value=(ar["모멘텀점수(가중합,%)"] if pd.notna(ar["모멘텀점수(가중합,%)"]) else None))
        tot.number_format = "0.00"; tot.border = border_all
        row_cursor += 2

    for grp in ["공격자산","안전자산"]:
        sub = df[df["분류"]==grp]
        if sub.empty: continue
        ws.cell(row=row_cursor, column=1, value=grp).font = Font(size=12, bold=True)
        row_cursor += 1
        for _, ar in sub.iterrows():
            add_block(ar)

    ws.freeze_panes = "A5"
    autosize_columns(ws, max_width=60)

# =========================================================
# 메인
# =========================================================
def main():
    month_str = datetime.now().strftime("%Y-%m")
    os.makedirs(OUT_DIR, exist_ok=True)
    xlsx_path = os.path.join(OUT_DIR, f"vaa_composite_report_{month_str}.xlsx")

    summary = build_summary_df()
    banner  = decision_banner(summary)

    wb = Workbook(); wb.remove(wb.active)
    write_summary_sheet(wb, summary, month_str)
    write_detail_sheet(wb, summary, banner, month_str)
    wb.save(xlsx_path)

    print(f"✅ 엑셀 저장 완료: {xlsx_path}")
    print(f"📌 {banner}")

if __name__ == "__main__":
    main()