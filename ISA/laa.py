# laa_strategy_report.py
# -*- coding: utf-8 -*-
import os
from datetime import datetime
import numpy as np
import pandas as pd
import yfinance as yf
from pandas_datareader import data as pdr

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

OUT_DIR = "laa_out"
os.makedirs(OUT_DIR, exist_ok=True)

# =========================
# 설정
# =========================
START_DATE = "2000-01-01"
SPX_TICKER = "^GSPC"       # S&P 500 Index
UNRATE_SER = "UNRATE"      # FRED 미국 실업률(%) 월간
FIXED_ASSETS = [
    ("미국 대형가치주", "IWD", 0.25),
    ("금",           "GLD", 0.25),
    ("미국 중기국채", "IEF", 0.25),
]
TIMING_PAIR = ("QQQ", "SHY")  # (위험자산, 안전자산)
TIMING_WEIGHT = 0.25

# =========================
# 데이터 로딩
# =========================
def load_daily_close(ticker: str, start=START_DATE) -> pd.Series:
    # auto_adjust를 명시적으로 False로 지정해 경고 제거
    df = yf.download(ticker, start=start, progress=False, auto_adjust=False)
    if df.empty or "Close" not in df:
        raise RuntimeError(f"{ticker} 데이터를 불러오지 못했습니다.")
    s = df["Close"]
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:, 0]
    s = s.dropna()
    s.name = ticker
    return s

def load_unrate(start=START_DATE) -> pd.Series:
    """FRED UNRATE(%) 월간 → 반드시 1D Series로 정규화"""
    df = pdr.DataReader(UNRATE_SER, "fred", start=start)
    if df is None or df.empty:
        raise RuntimeError("UNRATE 데이터를 불러오지 못했습니다.")
    # squeeze로 1D 보장
    s = df.squeeze("columns")
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:, 0]
    s = s.dropna()
    s.name = "UNRATE(%)"
    return s

# =========================
# 신호 계산
# =========================
def compute_signals():
    # 1) S&P500 일별 종가 + 200거래일 SMA
    spx = load_daily_close(SPX_TICKER)
    spx_sma200 = spx.rolling(window=200, min_periods=200).mean()

    # 2) 월말 샘플링: 'ME' 사용 + 월(period) 정렬
    #    - 거래일 월말이 실제 달의 말일과 다를 수 있으므로 period 기준으로 정규화
    spx_me = spx.resample("ME").last()  # month end
    sma200_me = spx_sma200.reindex(spx.index, method="ffill").resample("ME").last()

    # period(M) 인덱스로 변환
    spx_p = spx_me.to_period("M")
    sma200_p = sma200_me.to_period("M")

    # 3) UNRATE 월간 + 12개월 이동평균 (원래 월간이므로 바로 period로 맞춤)
    ur = load_unrate()
    ur_12m = ur.rolling(window=12, min_periods=12).mean()

    ur_p = ur.to_period("M")
    ur12_p = ur_12m.to_period("M")

    # 4) 공통 period(M) 인덱스 교집합
    common_p = spx_p.index.intersection(ur_p.index)
    if len(common_p) == 0:
        raise RuntimeError("공통 월(period) 인덱스가 비었습니다. 데이터 수집 기간/네트워크를 확인하세요.")

    # 5) 공통 period로 재색인 후 month-end 타임스탬프로 되돌리기
    spx_p = spx_p.reindex(common_p)
    sma200_p = sma200_p.reindex(common_p)
    ur_p = ur_p.reindex(common_p)
    ur12_p = ur12_p.reindex(common_p)

    # period → 실제 달의 말일 타임스탬프
    idx_ts = common_p.to_timestamp("M", how="end")

    spx_m = pd.Series(spx_p.values, index=idx_ts, name="SPX_Close")
    sma200_m = pd.Series(sma200_p.values, index=idx_ts, name="SPX_200D_SMA")
    ur_m = pd.Series(ur_p.values, index=idx_ts, name="UNRATE(%)")
    ur12_m = pd.Series(ur12_p.values, index=idx_ts, name="UNRATE_12M(%)")

    # 6) 합치기
    df = pd.concat([spx_m, sma200_m, ur_m, ur12_m], axis=1)

    # 7) 신호 계산
    cond_price = df["SPX_Close"] < df["SPX_200D_SMA"]
    cond_unemp = df["UNRATE(%)"] > df["UNRATE_12M(%)"]
    df["TimingChoice"] = np.where(cond_price & cond_unemp, "SHY", "QQQ")

    # 초기 결측 제거
    df = df.dropna(subset=["SPX_Close", "SPX_200D_SMA", "UNRATE(%)", "UNRATE_12M(%)"])

    return df

# =========================
# 현재 배분(목표 비중) 산출
# =========================
def current_allocation(signals_df: pd.DataFrame):
    if signals_df.empty:
        raise RuntimeError("신호 데이터가 비어 있습니다.")
    last_date = signals_df.index[-1]
    timing_choice = signals_df.loc[last_date, "TimingChoice"]

    # 고정자산(연 1회 리밸런싱 가정의 '목표 비중'을 보여줌)
    rows = []
    for name, ticker, w in FIXED_ASSETS:
        rows.append([name, ticker, round(w * 100, 2)])

    # 타이밍(월 1회 리밸런싱)
    rows.append(["타이밍 자산", timing_choice, round(TIMING_WEIGHT * 100, 2)])

    alloc_df = pd.DataFrame(rows, columns=["자산군", "티커", "목표비중(%)"])
    alloc_df["기준일"] = last_date.date()
    return alloc_df, last_date, timing_choice

# =========================
# 엑셀 저장
# =========================
def autosize_columns(ws, max_width=60):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            v = "" if v is None else str(v)
            widths[i] = max(widths.get(i, 0), len(v))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 12), max_width)

def build_excel(signals_df: pd.DataFrame, alloc_df: pd.DataFrame, last_date, timing_choice):
    month_str = datetime.now().strftime("%Y-%m")
    xlsx_path = os.path.join(OUT_DIR, f"laa_report_{month_str}.xlsx")

    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    # 공통 스타일
    title_fill = PatternFill("solid", fgColor="E6F0FF")
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="D9D9D9")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Named styles
    if "percent_style" not in wb.named_styles:
        st = NamedStyle(name="percent_style"); st.number_format = "0.00%"; wb.add_named_style(st)
    if "number_style" not in wb.named_styles:
        st = NamedStyle(name="number_style"); st.number_format = "#,##0.00"; wb.add_named_style(st)

    # ===== Sheet 1: Summary (현재 목표 배분) =====
    ws1 = wb.create_sheet("Summary")
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c = ws1.cell(row=1, column=1, value=f"LAA Summary — {month_str}")
    c.font = Font(size=14, bold=True); c.fill = title_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 24

    # 배너
    banner = f"현재 타이밍 자산: {timing_choice}  |  기준일: {last_date.date()}"
    ws1.cell(row=3, column=1, value=banner).font = Font(bold=True)

    # 표 헤더
    headers = ["자산군", "티커", "목표비중(%)", "비고"]
    for col, h in enumerate(headers, start=1):
        cell = ws1.cell(row=5, column=col, value=h)
        cell.font = Font(bold=True); cell.fill = header_fill
        cell.border = border_all; cell.alignment = Alignment(horizontal="center")

    # 데이터
    r = 6
    for _, row in alloc_df.iterrows():
        ws1.cell(row=r, column=1, value=row["자산군"]).border = border_all
        ws1.cell(row=r, column=2, value=row["티커"]).border = border_all
        c3 = ws1.cell(row=r, column=3, value=float(row["목표비중(%)"]) / 100.0)
        c3.border = border_all; c3.style = "percent_style"
        ws1.cell(row=r, column=4, value="" if row["자산군"] != "타이밍 자산" else "월 1회 리밸런싱").border = border_all
        r += 1

    ws1.freeze_panes = "A6"
    autosize_columns(ws1, max_width=40)

    # ===== Sheet 2: Signals (월말 S&P/200D, 실업률/12M, 타이밍) =====
    ws2 = wb.create_sheet("Signals")

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    c2 = ws2.cell(row=1, column=1, value=f"Signals — 월말 기준 (S&P500 vs 200D, 실업률 vs 12M, 타이밍) — {month_str}")
    c2.font = Font(size=14, bold=True); c2.fill = title_fill
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    # 보기 좋게 최근 120개월만 표시 (필요시 변경)
    sig = signals_df.copy()
    if len(sig) > 120:
        sig = sig.iloc[-120:].copy()

    sig_out = sig.reset_index().rename(columns={
        "index": "월말",
        "SPX_Close": "미국 S&P 500 지수 가격",
        "SPX_200D_SMA": "200일 이동평균 가격",
        "UNRATE(%)": "미국 실업률(%)",
        "UNRATE_12M(%)": "12개월 이동평균(%)",
        "TimingChoice": "타이밍 선택(월말)"
    })

    # 헤더
    for col, h in enumerate(sig_out.columns, start=1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True); cell.fill = header_fill
        cell.border = border_all; cell.alignment = Alignment(horizontal="center")

    # 데이터 + 서식
    r = 4
    for _, row in sig_out.iterrows():
        ws2.cell(row=r, column=1, value=row["월말"].date()).border = border_all

        cpx = ws2.cell(row=r, column=2, value=float(row["미국 S&P 500 지수 가격"]))
        cpx.border = border_all; cpx.style = "number_style"

        csma = ws2.cell(row=r, column=3, value=float(row["200일 이동평균 가격"]))
        csma.border = border_all; csma.style = "number_style"

        # 실업률은 백분율 → 엑셀 퍼센트 서식 적용
        u = row["미국 실업률(%)"]
        u12 = row["12개월 이동평균(%)"]
        cu = ws2.cell(row=r, column=4, value=None if pd.isna(u) else float(u) / 100.0)
        cu.border = border_all; cu.style = "percent_style"
        cu12 = ws2.cell(row=r, column=5, value=None if pd.isna(u12) else float(u12) / 100.0)
        cu12.border = border_all; cu12.style = "percent_style"

        ws2.cell(row=r, column=6, value=row["타이밍 선택(월말)"]).border = border_all

        r += 1

    ws2.freeze_panes = "A4"
    autosize_columns(ws2, max_width=52)

    # ===== Sheet 3: TimingOnly (월말 타이밍만 모아서) =====
    ws3 = wb.create_sheet("TimingOnly")
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    c3 = ws3.cell(row=1, column=1, value=f"Timing Choice History — {month_str}")
    c3.font = Font(size=14, bold=True); c3.fill = title_fill
    c3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 24

    # 최근 120개월만
    t_only = signals_df[["TimingChoice"]].copy()
    if len(t_only) > 120:
        t_only = t_only.iloc[-120:].copy()
    t_only = t_only.reset_index().rename(columns={"index": "월말", "TimingChoice": "타이밍 선택(월말)"})

    for col, h in enumerate(t_only.columns, start=1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True); cell.fill = header_fill
        cell.border = border_all; cell.alignment = Alignment(horizontal="center")

    r = 4
    for _, row in t_only.iterrows():
        ws3.cell(row=r, column=1, value=row["월말"].date()).border = border_all
        ws3.cell(row=r, column=2, value=row["타이밍 선택(월말)"]).border = border_all
        r += 1

    ws3.freeze_panes = "A4"
    autosize_columns(ws3, max_width=32)

    # 저장
    wb.save(xlsx_path)
    print(f"✅ 엑셀 저장 완료: {xlsx_path}")

# =========================
# main
# =========================
if __name__ == "__main__":
    # 1) 신호 테이블 생성
    signals = compute_signals()
    # 2) 현재 목표 배분 생성
    alloc, last_dt, timing = current_allocation(signals)
    # 3) 엑셀 출력 (Summary / Signals / TimingOnly)
    build_excel(signals, alloc, last_dt, timing)
    print(f"📌 현재 타이밍 자산: {timing} (기준일: {last_dt.date()})")
