# -*- coding: utf-8 -*-
# dualmomentom_isa_alternative3.py
# -----------------------------------------------------------------------------
# [대안 3] 완벽 일치형 듀얼모멘텀 (ISA 계좌 전용)
# - 의사결정: SPY(미국) vs [TIGER 유로스탁스50 + TIGER 일본니케이225 합성](선진국)
# - 실행: 국내 상장 해외 ETF (ISA 거래 가능)
# -----------------------------------------------------------------------------

import os
from datetime import datetime
import numpy as np
import pandas as pd
import yfinance as yf

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# 1. 설정 (ISA 포트폴리오)
# =========================
OUT_DIR = "dual_momentum_isa"
os.makedirs(OUT_DIR, exist_ok=True)

# 1) 의사결정용 티커 (야후 파이낸스 기준)
# - 미국 대표: SPY (데이터 역사가 길어서 판단용으로 적합)
# - 현금/채권: BIL (초단기채, 수비 기준)
# - 선진국(비미국): EFA 대신 실제 투자할 '국내 ETF' 데이터를 직접 사용 (괴리 제거)
TICKER_DECISION = {
    "US": "SPY",            # 미국 주식 판단
    "CASH": "BIL",          # 현금성 자산 판단 (절대모멘텀 기준)
    "EU_ETF": "195930.KS",  # TIGER 유로스탁스50(합성 H)
    "JP_ETF": "241180.KS"   # TIGER 일본니케이225
}

# 2) 실제 매수할 종목 (ISA 계좌용 국내상장 ETF)
# - 당첨된 자산군에 따라 매수할 종목 리스트
ALLOCATION_MAP = {
    "US_WIN": [
        {"지역": "미국", "종목명": "TIGER 미국S&P500", "Code": "360750", "비중": 1.0}
    ],
    "NON_US_WIN": [
        {"지역": "유럽", "종목명": "TIGER 유로스탁스50(합성 H)", "Code": "195930", "비중": 0.5},
        {"지역": "일본", "종목명": "TIGER 일본니케이225", "Code": "241180", "비중": 0.5}
    ],
    "DEFENSIVE": [
        {"지역": "채권", "종목명": "KODEX 미국종합채권SRI액티브(H)", "Code": "437080", "비중": 1.0}
    ]
}

# =========================
# 2. 데이터 유틸리티
# =========================
def get_monthly_close(ticker, start="2015-01-01"):
    """야후 파이낸스에서 월말 수정종가(Adj Close) 가져오기"""
    try:
        df = yf.download(ticker, start=start, progress=False, auto_adjust=True)
        if df.empty:
            return pd.Series(dtype=float)
        
        # 'Close' 컬럼 추출 (MultiIndex 처리)
        if isinstance(df.columns, pd.MultiIndex):
            # yfinance 최신 버전 대응
            try:
                s = df["Close"][ticker]
            except KeyError:
                s = df.iloc[:, 0] # 첫번째 컬럼 강제 선택
        else:
            s = df["Close"]
            
        # 월말 리샘플링
        monthly = s.resample("M").last().dropna()
        return monthly
    except Exception as e:
        print(f"Error fetching {ticker}: {e}")
        return pd.Series(dtype=float)

def calc_12m_return(monthly_series):
    """최근 12개월 수익률 계산 (현재 월말 / 12개월 전 월말 - 1)"""
    if len(monthly_series) < 13:
        return None
    p_now = float(monthly_series.iloc[-1])
    p_prev = float(monthly_series.iloc[-13])
    return (p_now / p_prev) - 1.0

# =========================
# 3. 핵심 로직: [대안 3] 적용
# =========================
def run_dual_momentum_alt3():
    print(">>> 데이터 수집 중...")
    
    # 1) 데이터 가져오기
    m_spy = get_monthly_close(TICKER_DECISION["US"])
    m_bil = get_monthly_close(TICKER_DECISION["CASH"])
    m_eu  = get_monthly_close(TICKER_DECISION["EU_ETF"])
    m_jp  = get_monthly_close(TICKER_DECISION["JP_ETF"])

    # 2) '합성 선진국 지수' 만들기 (유로50 + 니케이225 반반)
    # - 날짜 인덱스 맞추기 (교집합)
    idx = m_eu.index.intersection(m_jp.index)
    if len(idx) < 13:
        raise ValueError("국내 ETF 데이터가 부족하여 12개월 모멘텀을 계산할 수 없습니다. (상장일 확인 필요)")
    
    m_eu = m_eu.loc[idx]
    m_jp = m_jp.loc[idx]
    
    # - 월간 수익률 계산
    r_eu = m_eu.pct_change().fillna(0)
    r_jp = m_jp.pct_change().fillna(0)
    
    # - 합성 수익률 (50:50 리밸런싱 가정)
    r_composite = (r_eu * 0.5) + (r_jp * 0.5)
    
    # - 합성 지수화 (기준일 1.0 시작)
    #   (1+r).cumprod()를 통해 12개월 수익률 계산용 가상의 가격(Index) 생성
    m_composite_idx = (1 + r_composite).cumprod()

    # 3) 12개월 모멘텀 계산
    mom_spy = calc_12m_return(m_spy)
    mom_bil = calc_12m_return(m_bil)
    mom_composite = calc_12m_return(m_composite_idx) # 우리가 만든 합성 지수의 12개월 수익률

    if any(x is None for x in [mom_spy, mom_bil, mom_composite]):
        raise ValueError("최근 12개월 데이터가 부족합니다.")

    # 4) 듀얼모멘텀 판정 로직
    # Rule 1: 공격자산(SPY)이 안전자산(BIL)보다 좋은가? (절대모멘텀)
    #         * SPY 대신 Composite가 더 좋으면 Composite로도 비교해야 하나,
    #           전통 듀얼모멘텀은 보통 SPY를 기준으로 Market Stress를 판단하기도 함.
    #           여기서는 [SPY vs BIL] 비교 후, 공격 모드면 [SPY vs Composite] 승자를 고름.
    #           (단, 승자가 마이너스 모멘텀이면 BIL로 가는 로직도 추가 가능. 여기서는 Gary Antonacci 오리지널에 가깝게 SPY>BIL이면 공격으로 간주)
    
    decision_log = []
    final_choice = ""
    
    # 절대 모멘텀 체크 (SPY가 현금보다 강한가?)
    # *보수적 변형: SPY와 Composite 둘 중 이기는 놈이 BIL보다 커야 한다.
    winner_mom = max(mom_spy, mom_composite)
    winner_name = "SPY" if mom_spy >= mom_composite else "Composite(EU+JP)"
    
    decision_log.append(f"1. 각 자산 12개월 수익률")
    decision_log.append(f"   - SPY (미국): {mom_spy:.2%}")
    decision_log.append(f"   - 합성 (유로+니케이): {mom_composite:.2%}")
    decision_log.append(f"   - BIL (초단기채): {mom_bil:.2%}")
    
    if winner_mom > mom_bil:
        # 공격 자산 매수
        if winner_name == "SPY":
            final_choice = "US_WIN"
            reason = f"공격모드 ON: SPY({mom_spy:.2%})가 합성({mom_composite:.2%}) 및 BIL보다 우위"
        else:
            final_choice = "NON_US_WIN"
            reason = f"공격모드 ON: 합성({mom_composite:.2%})이 SPY({mom_spy:.2%}) 및 BIL보다 우위"
    else:
        # 수비 자산 매수
        final_choice = "DEFENSIVE"
        reason = f"수비모드 ON: 1등({winner_name}, {winner_mom:.2%})이 BIL({mom_bil:.2%})보다 낮음"

    print(f"\n[판정 결과] {reason}")
    
    return {
        "mom_spy": mom_spy,
        "mom_composite": mom_composite,
        "mom_bil": mom_bil,
        "final_choice": final_choice,
        "reason": reason,
        "m_composite_idx": m_composite_idx  # 차트/기록용
    }

# =========================
# 4. 엑셀 리포트 생성
# =========================
def save_report_to_excel(res_data):
    month_str = datetime.now().strftime("%Y-%m")
    filename = f"DualMomentum_ISA_Alt3_{month_str}.xlsx"
    filepath = os.path.join(OUT_DIR, filename)

    wb = Workbook()
    
    # 스타일 정의
    title_font = Font(size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="4472C4") # 파란색 헤더
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    center_align = Alignment(horizontal="center", vertical="center")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Sheet 1: 투자 리포트 ---
    ws = wb.active
    ws.title = "ISA 투자지시서"
    
    # 1. 제목
    ws.merge_cells("A1:E1")
    ws["A1"] = f"ISA 듀얼모멘텀 (대안3: 완전일치형) - {month_str}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center_align
    
    # 2. 이번 달 결정
    ws["A3"] = "결정 내역"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = res_data["reason"]
    
    # 3. 모멘텀 비교표
    headers = ["자산군", "티커(Data)", "12개월 수익률", "비고"]
    data_rows = [
        ["미국 주식", TICKER_DECISION["US"], res_data["mom_spy"], "S&P500 기준"],
        ["선진국(비미국)", "합성(195930+241180)", res_data["mom_composite"], "유로50+니케이225 (5:5)"],
        ["현금/채권", TICKER_DECISION["CASH"], res_data["mom_bil"], "Risk Free 기준"]
    ]
    
    # 표 헤더
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.fill = header_fill
        c.font = Font(bold=True)
        c.alignment = center_align
        c.border = border_thin

    # 표 내용
    for i, row in enumerate(data_rows, 6):
        ws.cell(row=i, column=1, value=row[0]).border = border_thin
        ws.cell(row=i, column=2, value=row[1]).border = border_thin
        ws.cell(row=i, column=3, value=row[2]).number_format = '0.00%'
        ws.cell(row=i, column=3).border = border_thin
        ws.cell(row=i, column=4, value=row[3]).border = border_thin
        
        # 승자 강조 (Bold + Color)
        val = row[2]
        if val == max(res_data["mom_spy"], res_data["mom_composite"], res_data["mom_bil"]):
             ws.cell(row=i, column=3).font = Font(bold=True, color="FF0000")

    # 4. 실제 매수 포트폴리오 (Allocation)
    ws["A9"] = "📢 이번 달 매수 종목 (ISA 계좌)"
    ws["A9"].font = Font(bold=True, size=12)
    
    alloc_headers = ["구분", "종목명", "종목코드", "투자비중"]
    for col, h in enumerate(alloc_headers, 1):
        c = ws.cell(row=10, column=col, value=h)
        c.fill = header_fill
        c.font = Font(bold=True)
        c.alignment = center_align
        c.border = border_thin
        
    target_portfolio = ALLOCATION_MAP[res_data["final_choice"]]
    
    start_row = 11
    for item in target_portfolio:
        ws.cell(row=start_row, column=1, value=item["지역"]).border = border_thin
        ws.cell(row=start_row, column=2, value=item["종목명"]).border = border_thin
        ws.cell(row=start_row, column=3, value=item["Code"]).border = border_thin
        
        c_weight = ws.cell(row=start_row, column=4, value=item["비중"])
        c_weight.number_format = '0%'
        c_weight.border = border_thin
        c_weight.fill = PatternFill("solid", fgColor="FFF2CC") # 노란색 강조
        start_row += 1

    # 컬럼 너비 조정
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30

    wb.save(filepath)
    print(f"✅ 리포트 생성 완료: {filepath}")

# =========================
# Main Execution
# =========================
if __name__ == "__main__":
    try:
        # 1. 듀얼모멘텀 분석 실행
        result = run_dual_momentum_alt3()
        
        # 2. 결과 엑셀 저장
        save_report_to_excel(result)
        
    except Exception as e:
        print(f"❌ 실행 중 오류 발생: {e}")