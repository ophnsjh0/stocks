# -*- coding: utf-8 -*-
# vaa_correlation_check.py
# VAA 전략 자산군(미국 ETF) vs 국내 대체 ETF 상관관계 및 추적오차 분석 리포트

import os
from datetime import datetime
import pandas as pd
import numpy as np
import yfinance as yf
import FinanceDataReader as fdr
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================================================
# 1. 설정 및 매핑 (vaa.py 내용 기반)
# =========================================================
OUT_DIR = "vaa_analysis_out"
os.makedirs(OUT_DIR, exist_ok=True)

# vaa.py에 정의된 매핑 정보
# 환노출: 미국 ETF(USD) * 환율 = 국내 ETF(KRW) 움직임 예상
# 환해지: 미국 ETF(USD)       = 국내 ETF(KRW) 움직임 예상
ASSET_PAIRS = {
    "SPY": {"KR_Code": "379800", "KR_Name": "KODEX 미국S&P500",              "Hedge": False}, # 환노출
    "EFA": {"KR_Code": "251350", "KR_Name": "KODEX MSCI선진국",              "Hedge": False}, # 환노출
    "EEM": {"KR_Code": "195980", "KR_Name": "PLUS 신흥국MSCI(합성 H)",          "Hedge": True},  # 환해지
    "AGG": {"KR_Code": "468630", "KR_Name": "KODEX iShares미국투자등급회사채 엑티브", "Hedge": False}, # 환노출
    "LQD": {"KR_Code": "437080", "KR_Name": "KODEX 미국종합채권ESG엑티브(H)", "Hedge": True},  # 환해지
    "IEF": {"KR_Code": "0085P0", "KR_Name": "ACE 미국10년국채엑티브",         "Hedge": False}, # 환노출
    "SHY": {"KR_Code": "440650", "KR_Name": "ACE 미국달러단기채권엑티브",     "Hedge": False}, # 환노출
}

# =========================================================
# 2. 데이터 다운로드 유틸리티
# =========================================================
def get_data(ticker, start="2018-01-01"):
    """야후 파이낸스(미국) 또는 FDR(한국) 데이터 다운로드"""
    try:
        # 한국 ETF (숫자로 구성되거나 .KS 등)
        if ticker.isdigit() or ticker.endswith(".KS"):
            df = fdr.DataReader(ticker, start=start)
            if df is None or df.empty:
                return pd.Series(dtype=float)
            return df['Close']
        
        # 미국 ETF (문자열)
        else:
            df = yf.download(ticker, start=start, progress=False, auto_adjust=True)
            if df.empty:
                return pd.Series(dtype=float)
            
            # 멀티인덱스 컬럼 처리 (yfinance 최신 버전)
            if isinstance(df.columns, pd.MultiIndex):
                try:
                    return df["Close"][ticker]
                except KeyError:
                    return df.iloc[:, 0]
            return df["Close"]
    except Exception as e:
        print(f"❌ 데이터 다운로드 실패 ({ticker}): {e}")
        return pd.Series(dtype=float)

def get_usdkrw(start="2018-01-01"):
    """환율 데이터 (USD/KRW)"""
    try:
        df = fdr.DataReader("USD/KRW", start=start)
        return df['Close']
    except Exception as e:
        print(f"❌ 환율 데이터 실패: {e}")
        return pd.Series(dtype=float)

# =========================================================
# 3. 분석 로직
# =========================================================
def analyze_pair(us_ticker, kr_info, fx_series):
    print(f">> 분석 중: {us_ticker} vs {kr_info['KR_Name']} ({kr_info['KR_Code']})...")
    
    # 데이터 가져오기
    us_close = get_data(us_ticker)
    kr_close = get_data(kr_info['KR_Code'])
    
    # 데이터 공통 구간 맞추기
    common_idx = us_close.index.intersection(kr_close.index).intersection(fx_series.index)
    
    if len(common_idx) < 200: # 데이터가 너무 적으면 스킵
        print(f"   ⚠️ 데이터 부족으로 건너뜀 (공통 거래일 {len(common_idx)}일)")
        return None

    us_s = us_close.loc[common_idx]
    kr_s = kr_close.loc[common_idx]
    fx_s = fx_series.loc[common_idx]

    # 비교 기준 생성 (Benchmark)
    # 환노출(Unhedged)이면: 미국지수 * 환율
    # 환해지(Hedged)이면:   미국지수 그대로
    if kr_info['Hedge']:
        benchmark = us_s
        hedge_str = "환해지(H)"
    else:
        benchmark = us_s * fx_s
        hedge_str = "환노출(UH)"

    # 일간 수익률 계산
    bm_ret = benchmark.pct_change().dropna()
    kr_ret = kr_s.pct_change().dropna()

    # 다시 인덱스 정렬
    idx = bm_ret.index.intersection(kr_ret.index)
    bm_ret = bm_ret.loc[idx]
    kr_ret = kr_ret.loc[idx]

    # 월간 데이터로 변환 (상관계수 계산용)
    bm_monthly = (1 + bm_ret).resample('M').prod() - 1
    kr_monthly = (1 + kr_ret).resample('M').prod() - 1
    
    # 분석 지표 계산 함수
    def calc_metrics(window_months):
        if len(bm_monthly) < window_months:
            return np.nan, np.nan, np.nan
        
        b = bm_monthly.iloc[-window_months:]
        k = kr_monthly.iloc[-window_months:]
        
        # 상관계수
        corr = np.corrcoef(b, k)[0, 1]
        
        # 누적 수익률 차이
        cum_b = (1 + b).prod() - 1
        cum_k = (1 + k).prod() - 1
        diff_cum = cum_k - cum_b

        # 추적 오차 (Tracking Error, 연율화)
        # 일간 데이터 기준 계산
        window_days = int(window_months * 21) # 영업일 대략 환산
        if len(bm_ret) >= window_days:
            diff_daily = kr_ret.iloc[-window_days:] - bm_ret.iloc[-window_days:]
            te = diff_daily.std() * np.sqrt(252) * 100 # % 단위
        else:
            te = np.nan
            
        return corr, diff_cum * 100, te

    res = {
        "Ticker": us_ticker,
        "KR_Name": kr_info['KR_Name'],
        "KR_Code": kr_info['KR_Code'],
        "Type": hedge_str,
        "Data_Days": len(idx)
    }

    for m in [12, 24, 36]:
        c, d, t = calc_metrics(m)
        res[f"Corr_{m}M"] = c
        res[f"Diff_{m}M"] = d
        res[f"TE_{m}M"] = t

    return res

# =========================================================
# 4. 엑셀 저장
# =========================================================
def save_excel(results):
    month_str = datetime.now().strftime("%Y-%m")
    filename = f"VAA_Asset_Correlation_{month_str}.xlsx"
    filepath = os.path.join(OUT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "상관관계 분석"

    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = [
        "미국 자산", "국내 대체 자산", "종목코드", "환헤지 여부", 
        "12M 상관계수", "24M 상관계수", "36M 상관계수",
        "12M 추적오차(%)", "24M 추적오차(%)", "36M 추적오차(%)",
        "12M 수익률 괴리(%)", "데이터(일)"
    ]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center_align
        c.border = border

    # 데이터 입력
    for i, r in enumerate(results, 2):
        row_vals = [
            r["Ticker"], r["KR_Name"], r["KR_Code"], r["Type"],
            r["Corr_12M"], r["Corr_24M"], r["Corr_36M"],
            r["TE_12M"], r["TE_24M"], r["TE_36M"],
            r["Diff_12M"], r["Data_Days"]
        ]

        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.border = border
            c.alignment = center_align
            
            # 포맷팅
            if isinstance(val, float):
                if col in [5, 6, 7]: # 상관계수
                    c.number_format = "0.0000"
                    # 상관계수 조건부 서식 (0.9 미만 빨간색)
                    if val < 0.9: c.font = Font(color="FF0000", bold=True)
                else: # 오차, 괴리
                    c.number_format = "0.00"

    # 컬럼 너비 조정
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["D"].width = 12
    for c_idx in range(5, 12):
        ws.column_dimensions[get_column_letter(c_idx)].width = 15

    wb.save(filepath)
    print(f"\n✅ 리포트 생성 완료: {filepath}")

# =========================================================
# 메인 실행
# =========================================================
if __name__ == "__main__":
    print(">>> 환율 데이터 다운로드 중...")
    fx_data = get_usdkrw()
    
    if fx_data.empty:
        print("❌ 환율 데이터를 가져올 수 없어 종료합니다.")
    else:
        results = []
        for us_ticker, info in ASSET_PAIRS.items():
            res = analyze_pair(us_ticker, info, fx_data)
            if res:
                results.append(res)
        
        if results:
            save_excel(results)
        else:
            print("❌ 분석된 데이터가 없습니다.")