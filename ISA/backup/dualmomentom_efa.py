# -*- coding: utf-8 -*-
# dualmomentom_returns_report.py
# - 듀얼모멘텀(미국 ETF로 의사결정 → 국내 ETF로 실행)
# - EFA 매핑: KODEX MSCI선진국(251350) 단일
# - 비교 시트: EFA vs 251350 (12/24/36M 상관, 추적차, 거래량)

import os
from datetime import datetime
import numpy as np
import pandas as pd
import yfinance as yf

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# 기본 설정
# =========================
OUT_DIR = "dual_momentum_out"
os.makedirs(OUT_DIR, exist_ok=True)

# =========================
# 백데이터(의사결정) 티커
# =========================
US_TICKERS = {
    "SPY": "미국 주식SPY",
    "EFA": "선진국 주식EFA",
    "BIL": "초단기채권BIL",
    "AGG": "미국 혼합채권AGG",  # fallback
}

# =========================
# 실제 투자 매핑 (국내 ETF)
# =========================
KR_MAPPING = {
    "SPY": [
        {"분류": "미국",   "종목명": "KODEX 미국S&P500",                 "Code": "379800", "환율": "환해지", "비중(%)": 100.0},
    ],
    "EFA": [
        # ✅ 변경: EFA는 KODEX MSCI선진국(251350) 단일 매핑
        {"분류": "선진국", "종목명": "KODEX MSCI선진국",                 "Code": "251350", "환율": "환노출", "비중(%)": 100.0},
    ],
    "AGG": [
        {"분류": "채권",   "종목명": "KODEX 미국종합채권SRI액티브(H)",   "Code": "437080", "환율": "환노출", "비중(%)": 100.0},
    ],
}

# 국내 ETF 코드 목록 (Returns 시트 계산용)
KR_CODES = ["379800", "251350", "437080"]


# =========================
# 유틸
# =========================
def ensure_series(x: pd.Series | pd.DataFrame) -> pd.Series:
    """Close 등에서 뽑은 뒤에도 가끔 DataFrame이 남는 케이스 방지: 1열 Series 강제."""
    if isinstance(x, pd.DataFrame):
        return x.iloc[:, 0].astype(float)
    return x.astype(float)

def monthly_close(ticker: str, start="2010-01-01") -> pd.Series:
    """야후에서 받아 월말 종가 Series로 반환."""
    df = yf.download(ticker, start=start, progress=False)
    if df.empty or "Close" not in df:
        raise RuntimeError(f"{ticker} 데이터가 비어 있습니다.")
    m = df["Close"].resample("M").last().dropna()
    return ensure_series(m)

def trailing_12m_return(monthly: pd.Series) -> float:
    """최근 월말 기준 12개월 수익률 (비율, 0.1234=12.34%)."""
    if len(monthly) < 13:
        raise RuntimeError("12개월 수익률 계산에 필요한 월말 데이터가 부족합니다.")
    p0 = float(monthly.iloc[-1])     # 최근 월말
    p12 = float(monthly.iloc[-13])   # 12개월 전 월말
    return (p0 / p12) - 1.0


# =========================
# 의사결정 로직 (듀얼모멘텀)
# =========================
def decide_allocation():
    # 월말 시계열
    m_spy = monthly_close("SPY")
    m_efa = monthly_close("EFA")
    m_bil = monthly_close("BIL")
    m_agg = monthly_close("AGG")

    # 최근 12M 수익률
    r_spy = trailing_12m_return(m_spy)
    r_efa = trailing_12m_return(m_efa)
    r_bil = trailing_12m_return(m_bil)
    r_agg = trailing_12m_return(m_agg)

    # 룰:
    # 1) SPY 12M > BIL 12M → SPY vs EFA 중 12M 높은 ETF
    # 2) 아니면 AGG
    if r_spy > r_bil:
        chosen_us = "SPY" if r_spy >= r_efa else "EFA"
        rule_text = f"[룰1] SPY(12M={r_spy*100:.2f}%) > BIL(12M={r_bil*100:.2f}%) → SPY vs EFA 중 더 높은 12M → {chosen_us}"
    else:
        chosen_us = "AGG"
        rule_text = f"[룰2] SPY(12M={r_spy*100:.2f}%) ≤ BIL(12M={r_bil*100:.2f}%) → AGG 선택"

    # 실제 투자 배분표
    kr_alloc = pd.DataFrame(KR_MAPPING[chosen_us])

    # 요약표 (미국ETF 12M 수익률)
    summary = pd.DataFrame({
        "US_Ticker": ["SPY", "EFA", "BIL", "AGG"],
        "라벨": [US_TICKERS["SPY"], US_TICKERS["EFA"], US_TICKERS["BIL"], US_TICKERS["AGG"]],
        "12M수익률(%)": [round(r_spy*100, 2), round(r_efa*100, 2), round(r_bil*100, 2), round(r_agg*100, 2)]
    })

    # 배너 문구
    alloc_text = " + ".join([f"{row['종목명']}({row['Code']}) {row['비중(%)']:.0f}%" for _, row in kr_alloc.iterrows()])
    banner = f"이번달 실제 투자 대상: {alloc_text}  |  결정근거: {rule_text}"

    # 기준자산의 12M(%) 값 (Allocation 시트에 참고용으로 넣기)
    chosen_12m_pct = r_spy*100 if chosen_us == "SPY" else (r_efa*100 if chosen_us == "EFA" else r_agg*100)

    return summary, kr_alloc, banner, chosen_us, round(chosen_12m_pct, 2)


# =========================
# Returns 시트 데이터 (미국/국내 모두)
# =========================
def build_returns_sheet_data():
    rows = []

    # 미국 ETF 12M
    for t, label in US_TICKERS.items():
        try:
            r = trailing_12m_return(monthly_close(t)) * 100
            rows.append(["미국", label, t, None, None, round(r, 2)])
        except Exception:
            rows.append(["미국", label, t, None, None, None])

    # 국내 ETF 12M (야후 '.KS')
    for code in KR_CODES:
        y_ticker = f"{code}.KS"
        label = f"국내 ETF {code}"
        try:
            r = trailing_12m_return(monthly_close(y_ticker)) * 100
            rows.append(["국내", label, None, code, "KS", round(r, 2)])
        except Exception:
            rows.append(["국내", label, None, code, "KS", None])

    return pd.DataFrame(rows, columns=["구분","자산라벨","US_Ticker","KR_Code","시장","12M수익률(%)"])


# =========================
# 새 기능: EFA vs 251350 비교 데이터
# =========================
def build_compare_efa_vs_251350():
    """
    반환:
      metrics_df: 12/24/36M 상관계수 & 추적지표(누적차이, 평균월간차이, Tracking Error) & 최근 3개월 거래량 비교
      detail_df:  최근 36개월 월간 수익률(%) 시계열 비교표
    """
    # 월말 종가(Series 강제)
    m_efa = monthly_close("EFA")
    m_251 = monthly_close("251350.KS")

    # 공통 구간 정렬
    idx = m_efa.index.intersection(m_251.index)
    m_efa = m_efa.loc[idx].copy()
    m_251 = m_251.loc[idx].copy()

    # 월간 수익률(Series 강제)
    r_efa = ensure_series(m_efa.pct_change().dropna())
    r_251 = ensure_series(m_251.pct_change().dropna())
    ridx = r_efa.index.intersection(r_251.index)
    r_efa = r_efa.loc[ridx]
    r_251 = r_251.loc[ridx]

    def _window_slice(s: pd.Series, months: int):
        return s.iloc[-months:] if len(s) >= months else s.copy()

    def _cumret(x: pd.Series) -> float:
        return float((1.0 + x).prod() - 1.0) if len(x) else np.nan

    rows = []
    for win in [12, 24, 36]:
        re = _window_slice(r_efa, win)
        rk = _window_slice(r_251, win)
        # 길이/정렬 동일 보장
        ridx2 = re.index.intersection(rk.index)
        re = re.loc[ridx2]
        rk = rk.loc[ridx2]

        if len(re) > 2 and len(rk) == len(re):
            # Series.corr 대신 np.corrcoef로 안전 계산
            corr_val = float(np.corrcoef(re.values, rk.values)[0, 1])
            cum_diff = _cumret(rk) - _cumret(re)
            diff = (rk - re)
            mean_diff = float(diff.mean())
            te_monthly = float(diff.std(ddof=1)) if len(diff) > 2 else np.nan
            te_annual = te_monthly * np.sqrt(12) if np.isfinite(te_monthly) else np.nan
        else:
            corr_val = cum_diff = mean_diff = te_monthly = te_annual = np.nan

        rows.append([
            f"{win}M",
            None if not np.isfinite(corr_val) else round(corr_val, 4),
            None if not np.isfinite(cum_diff) else round(cum_diff * 100, 2),
            None if not np.isfinite(mean_diff) else round(mean_diff * 100, 3),
            None if not np.isfinite(te_monthly) else round(te_monthly * 100, 3),
            None if not np.isfinite(te_annual) else round(te_annual * 100, 3),
        ])

    metrics_df = pd.DataFrame(rows, columns=[
        "구간", "상관계수(월수익률)", "누적수익률 차이(국내−EFA, %)",
        "평균 월간 차이(%, 국내−EFA)", "Tracking Error(月, %)", "Tracking Error(연율, %)"
    ])

    # 거래량 비교(최근 3개월, 일봉)
    d_efa = yf.download("EFA", period="4mo", interval="1d", progress=False)
    d_251 = yf.download("251350.KS", period="4mo", interval="1d", progress=False)

    def last_n_months_mean_median_vol(df: pd.DataFrame, days=90):
        if df is None or df.empty or "Volume" not in df:
            return np.nan, np.nan
        last = df.tail(days)["Volume"].dropna()
        if last.empty:
            return np.nan, np.nan
        return float(last.mean()), float(last.median())

    efa_mean, efa_med = last_n_months_mean_median_vol(d_efa)
    k_mean, k_med   = last_n_months_mean_median_vol(d_251)

    vol_df = pd.DataFrame([
        ["EFA", efa_mean, efa_med],
        ["251350.KS", k_mean, k_med],
    ], columns=["티커", "최근3개월 일평균 거래량", "최근3개월 일중앙 거래량"])

    # 상세 시계열(최근 36개월 월수익률 %)
    r_join = pd.concat([
        ensure_series(r_efa).rename("EFA"),
        ensure_series(r_251).rename("251350.KS"),
    ], axis=1).dropna()
    r_detail = r_join.tail(36) * 100.0
    r_detail.index = r_detail.index.strftime("%Y-%m")

    return metrics_df, vol_df, r_detail.reset_index().rename(columns={"index": "월"})


# =========================
# 엑셀 저장
# =========================
def autosize_columns(ws, max_width=46):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            v = "" if v is None else str(v)
            widths[i] = max(widths.get(i, 0), len(v))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), max_width)

def save_excel(summary: pd.DataFrame, alloc: pd.DataFrame, banner: str, chosen_us: str, chosen_12m_pct: float,
               returns_df: pd.DataFrame, cmp_metrics: pd.DataFrame, cmp_vol: pd.DataFrame, cmp_detail: pd.DataFrame):
    month_str = datetime.now().strftime("%Y-%m")
    xlsx_path = os.path.join(OUT_DIR, f"dualmo_report_{month_str}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    title_fill = PatternFill("solid", fgColor="E6F0FF")
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="D9D9D9")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # === Sheet 1: Decision (미국ETF 12M 수익률 요약) ===
    ws1 = wb.create_sheet("Decision")
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c = ws1.cell(row=1, column=1, value=f"SPY/EFA/BIL 12M 모멘텀 의사결정 — {month_str}")
    c.font = Font(size=14, bold=True); c.fill = title_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 24

    ws1.cell(row=3, column=1, value=banner).font = Font(bold=True)

    for col, h in enumerate(summary.columns, start=1):
        cell = ws1.cell(row=5, column=col, value=h)
        cell.font = Font(bold=True); cell.fill = header_fill
        cell.border = border_all; cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(summary.itertuples(index=False), start=6):
        for c_idx, val in enumerate(row, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_all
            if summary.columns[c_idx-1].endswith("(%)") and isinstance(val, (int, float)):
                cell.number_format = "0.00%"; cell.value = val / 100.0

    ws1.freeze_panes = "A6"
    autosize_columns(ws1, max_width=36)

    # === Sheet 2: Allocation (실제 투자) ===
    ws2 = wb.create_sheet("Allocation")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    c2 = ws2.cell(row=1, column=1, value=f"실제 투자 배분 (국내 ETF) — {month_str}")
    c2.font = Font(size=14, bold=True); c2.fill = title_fill
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    headers2 = ["분류","종목명","Code","환율","비중(%)","(참고) 기준자산","(참고) 기준자산 12M(%)"]
    for col, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True); cell.fill = header_fill
        cell.border = border_all; cell.alignment = Alignment(horizontal="center")

    r_idx = 4
    for _, row in alloc.iterrows():
        ws2.cell(row=r_idx, column=1, value=row["분류"]).border = border_all
        ws2.cell(row=r_idx, column=2, value=row["종목명"]).border = border_all
        ws2.cell(row=r_idx, column=3, value=row["Code"]).border = border_all
        ws2.cell(row=r_idx, column=4, value=row["환율"]).border = border_all

        pct = float(row["비중(%)"]) / 100.0
        c = ws2.cell(row=r_idx, column=5, value=pct)
        c.border = border_all; c.number_format = "0.00%"

        ws2.cell(row=r_idx, column=6, value=US_TICKERS[chosen_us]).border = border_all

        c12 = ws2.cell(row=r_idx, column=7, value=chosen_12m_pct / 100.0)
        c12.border = border_all; c12.number_format = "0.00%"

        r_idx += 1

    ws2.freeze_panes = "A4"
    autosize_columns(ws2, max_width=46)

    # === Sheet 3: Returns (미국/국내 각 자산 12M 수익률) ===
    ws3 = wb.create_sheet("Returns")
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c3 = ws3.cell(row=1, column=1, value=f"각 자산 12개월 수익률 — {month_str}")
    c3.font = Font(size=14, bold=True); c3.fill = title_fill
    c3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 24

    for col, h in enumerate(returns_df.columns, start=1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True); cell.fill = header_fill
        cell.border = border_all; cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(returns_df.itertuples(index=False), start=4):
        for c_idx, val in enumerate(row, start=1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_all
            if returns_df.columns[c_idx-1].endswith("(%)") and isinstance(val, (int, float)):
                cell.number_format = "0.00%"; cell.value = val / 100.0

    autosize_columns(ws3, max_width=46)

    # === Sheet 4: Compare_EFA_vs_251350 ===
    ws4 = wb.create_sheet("Compare_EFA_vs_251350")
    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    c4 = ws4.cell(row=1, column=1, value=f"EFA vs KODEX MSCI선진국(251350) 비교 — {month_str}")
    c4.font = Font(size=14, bold=True); c4.fill = title_fill
    c4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 24

    # (A) 상관/추적 지표 표
    ws4.cell(row=3, column=1, value="A. 상관 & 추적지표").font = Font(bold=True)
    start_row = 5
    for col, h in enumerate(cmp_metrics.columns, start=1):
        cell = ws4.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True); cell.fill = header_fill
        cell.border = border_all; cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(cmp_metrics.itertuples(index=False), start=start_row+1):
        for c_idx, val in enumerate(row, start=1):
            ws4.cell(row=r_idx, column=c_idx, value=val).border = border_all

    # (B) 최근 3개월 거래량 표
    r2 = start_row + 1 + len(cmp_metrics) + 2
    ws4.cell(row=r2, column=1, value="B. 최근 3개월 일별 거래량(단순)").font = Font(bold=True)
    for col, h in enumerate(cmp_vol.columns, start=1):
        cell = ws4.cell(row=r2+2, column=col, value=h)
        cell.font = Font(bold=True); cell.fill = header_fill
        cell.border = border_all; cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(cmp_vol.itertuples(index=False), start=r2+3):
        for c_idx, val in enumerate(row, start=1):
            ws4.cell(row=r_idx, column=c_idx, value=val).border = border_all

    # (C) 최근 36개월 월수익률(%) 비교표
    r3 = r2 + 3 + len(cmp_vol) + 2
    ws4.cell(row=r3, column=1, value="C. 최근 36개월 월간 수익률(%)").font = Font(bold=True)
    for col, h in enumerate(cmp_detail.columns, start=1):
        cell = ws4.cell(row=r3+2, column=col, value=h)
        cell.font = Font(bold=True); cell.fill = header_fill
        cell.border = border_all; cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(cmp_detail.itertuples(index=False), start=r3+3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_all
            if c_idx >= 2 and isinstance(val, (int, float)):
                cell.number_format = "0.00"

    autosize_columns(ws4, max_width=52)

    # 저장
    wb.save(xlsx_path)
    print(f"✅ 엑셀 저장 완료: {xlsx_path}")


# =========================
# main
# =========================
if __name__ == "__main__":
    # 듀얼모멘텀 의사결정 및 기본 시트
    summary_df, alloc_df, banner_txt, chosen_us, chosen_12m_pct = decide_allocation()
    returns_df = build_returns_sheet_data()

    # 비교 시트(EFA vs 251350)
    cmp_metrics_df, cmp_vol_df, cmp_detail_df = build_compare_efa_vs_251350()

    # 엑셀 저장
    save_excel(summary_df, alloc_df, banner_txt, chosen_us, chosen_12m_pct,
               returns_df, cmp_metrics_df, cmp_vol_df, cmp_detail_df)

    print("📌", banner_txt)
